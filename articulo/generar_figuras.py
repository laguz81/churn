#!/usr/bin/env python3
"""
generar_figuras.py
Genera las tres figuras del artículo a partir de los respaldos de cálculo.
Ningún valor está escrito a mano: todo se lee de los xlsx.

Uso:
    python generar_figuras.py \
        --calculos  Respaldo_Calculos_v18m_20260811.xlsx \
        --ciego     Respaldo_Estudio_Ciego_20260818.xlsx \
        --salida    figuras/

Salidas: figura1_frontera.png, figura2_segmentos.png, figura3_diferencias.png
"""
from __future__ import annotations
import argparse, hashlib
from pathlib import Path

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
import openpyxl

AZUL, GRIS, ROJO = '#1f4e9c', '#b5b5b5', '#b03030'

plt.rcParams.update({
    'font.family': 'serif', 'font.size': 9,
    'axes.linewidth': .7, 'axes.edgecolor': '#333',
    'axes.grid': True, 'grid.color': '#e0e0e0', 'grid.linewidth': .5,
    'figure.dpi': 300,
})


def sha256(p: Path) -> str:
    h = hashlib.sha256()
    with open(p, 'rb') as f:
        for b in iter(lambda: f.read(65536), b''):
            h.update(b)
    return h.hexdigest()


def hoja(wb, nombre):
    ws = wb[nombre]
    filas = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    cab = next(i for i, r in enumerate(filas) if r[0] == 'regla' or 'segmento' == r[0])
    cols = [str(c) for c in filas[cab]]
    return cols, filas[cab + 1:]


# ── FIGURA 1 ────────────────────────────────────────────────────────
def figura1(wb, salida):
    cols, datos = hoja(wb, '11_Comparacion_Reglas')
    ic, im, ir = cols.index('contactos'), cols.index('margen_en_riesgo'), cols.index('regla')
    reglas = [(str(f[ir]), float(f[ic]), float(f[im])) for f in datos if f[ic] is not None]

    r = np.linspace(0, 25, 600)
    curvas = {n: (m - c * r) / 1000 for n, c, m in reglas}
    env = np.max(list(curvas.values()), axis=0)
    # una regla está dominada si nunca toca la envolvente
    dominada = {n: not np.any(np.isclose(v, env, atol=1e-9)) for n, v in curvas.items()}

    # cortes exactos: intersección de las dos rectas que se relevan
    # beneficio(r) = m - c*r  ->  cruce entre i y j en r = (m_i - m_j) / (c_i - c_j)
    def gana(x):
        return max(reglas, key=lambda t: t[2] - t[1] * x)[0]

    cand = sorted({(mi - mj) / (ci - cj)
                   for _, ci, mi in reglas for _, cj, mj in reglas
                   if ci != cj and 0 < (mi - mj) / (ci - cj) < 25})
    cortes, tramos, ini = [], [], 0.0
    for x in cand:
        e = 1e-6
        if gana(x - e) != gana(x + e):
            cortes.append(x)
            tramos.append((ini, x, gana(x - e)))
            ini = x
    tramos.append((ini, 25.0, gana(25.0 - 1e-6)))

    bonito = {'fijo_w=0.5': 'Fija 0,5 sem', 'fijo_w=1': 'Fija 1 sem', 'fijo_w=2': 'Fija 2 sem',
              'fijo_w=3': 'Fija 3 sem', 'fijo_w=4': 'Fija 4 sem', 'relativo_k=1.5': 'Relativa k=1,5',
              'relativo_k=2.0': 'Relativa k=2,0', 'relativo_k=2.5': 'Relativa k=2,5',
              'relativo_k=3.0': 'Relativa k=3,0'}

    fig, ax = plt.subplots(figsize=(6.8, 4.2))
    for k, (a, b, g) in enumerate(tramos):
        if k % 2 == 0:
            ax.axvspan(a, b, color='#f2f5fa', zorder=0)
        ax.text((a + b) / 2, 16.35, bonito.get(g, g), ha='center', va='center',
                fontsize=6.3, color=AZUL,
                bbox=dict(boxstyle='round,pad=0.22', fc='white', ec=AZUL, lw=.5))
    for n, c, m in reglas:
        ax.plot(r, curvas[n], color=GRIS if dominada[n] else AZUL,
                lw=.8 if dominada[n] else 1.3, ls=':' if dominada[n] else '-',
                zorder=2 if dominada[n] else 3)
    ax.plot(r, env, color='#c23b22', lw=3.2, alpha=.30, zorder=4, solid_capstyle='round')
    for x in cortes:
        ax.axvline(x, color='#777', lw=.6, ls='--', alpha=.7, zorder=1)
        ax.text(x, -0.75, f'{x:.2f}'.replace('.', ','), ha='center', fontsize=6.3, color='#555')
    ax.text(22.5, 1.0, 'reglas dominadas', fontsize=6.8, color='#8a8a8a',
            style='italic', ha='right')
    ax.set_xlim(0, 25); ax.set_ylim(0, 17.4); ax.set_yticks(range(0, 17, 4))
    ax.set_xlabel('Razón entre costo de contacto y probabilidad de retención  ($r$)', labelpad=12)
    ax.set_ylabel('Beneficio recuperable por unidad de\nprobabilidad de retención (miles de USD)')
    fig.tight_layout()
    fig.savefig(salida / 'figura1_frontera.png', bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f'  figura1  cortes={[round(c,2) for c in cortes]}  '
          f'dominadas={[n for n,v in dominada.items() if v]}')


# ── FIGURA 2 ────────────────────────────────────────────────────────
def figura2(wb, salida):
    ws = wb['13b_Desglose_TP_FP_Segmento']
    filas = [r for r in ws.iter_rows(values_only=True)
             if len(r) > 5 and r[1] and 'k=3.0' in str(r[1])]
    segs = [(str(f[0]), float(f[4]), float(f[5])) for f in filas]
    segs.sort(key=lambda x: -x[2])
    nom = [s[0].replace(' / ', ' /\n') for s in segs]
    cont = np.array([s[1] for s in segs]); marg = np.array([s[2] for s in segs])
    pc, pm, pcont = 100 * cont / cont.sum(), 100 * marg / marg.sum(), marg / cont

    fig, (a1, a2) = plt.subplots(1, 2, figsize=(6.8, 3.3))
    x, w = np.arange(len(segs)), .36
    a1.bar(x - w / 2, pc, w, label='Contactos', color=GRIS, zorder=3)
    a1.bar(x + w / 2, pm, w, label='Margen recuperable', color=AZUL, zorder=3)
    for i in range(len(segs)):
        a1.text(x[i] - w / 2, pc[i] + 1.6, f'{pc[i]:.1f}'.replace('.', ',') + ' %',
                ha='center', fontsize=6.6, color='#555')
        a1.text(x[i] + w / 2, pm[i] + 1.6, f'{pm[i]:.1f}'.replace('.', ',') + ' %',
                ha='center', fontsize=6.6, color=AZUL)
    a1.set_xticks(x); a1.set_xticklabels(nom, fontsize=7)
    a1.set_ylabel('Porcentaje del total', fontsize=8); a1.set_ylim(0, 100)
    a1.legend(fontsize=6.8, frameon=False, loc='upper right'); a1.grid(axis='x', visible=False)

    a2.bar(x, pcont, .5, color=AZUL, zorder=3)
    for i in range(len(segs)):
        a2.text(x[i], pcont[i] + 2.2, f'{pcont[i]:.2f}'.replace('.', ',') + ' USD',
                ha='center', fontsize=6.6, color=AZUL)
    a2.set_xticks(x); a2.set_xticklabels(nom, fontsize=7)
    a2.set_ylabel('Margen recuperable por contacto (USD)', fontsize=8)
    a2.set_ylim(0, pcont.max() * 1.22); a2.grid(axis='x', visible=False)
    fig.tight_layout()
    fig.savefig(salida / 'figura2_segmentos.png', bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f'  figura2  segmentos={[s[0] for s in segs]}')


# ── FIGURA 3 ────────────────────────────────────────────────────────
def figura3(wb, salida):
    ws = wb['03_Tabla_Pareada']
    filas = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    cab = next(i for i, f in enumerate(filas)
               if f and any('caso' == str(c).lower() for c in f if c))
    cols = [str(c).lower() if c else '' for c in filas[cab]]
    def col(*claves):
        for i, c in enumerate(cols):
            if all(k in c for k in claves):
                return i
        raise KeyError(claves)
    casos, rel, via = [], [], []
    for f in filas[cab + 1:]:
        if f[col('caso')] is None:
            continue
        casos.append(int(f[col('caso')]))
        rel.append(float(f[col('sistema', 'relev')]) - float(f[col('eh2', 'relev')]))
        via.append(float(f[col('sistema', 'viab')]) - float(f[col('eh2', 'viab')]))

    fig, axes = plt.subplots(1, 2, figsize=(6.9, 3.6))
    for ax, dif, tit in zip(axes, [rel, via], ['Relevancia', 'Viabilidad']):
        o = np.argsort(dif); y = np.arange(len(dif))
        v = [dif[i] for i in o]
        cols_ = [AZUL if k > 0 else (ROJO if k < 0 else '#c4c4c4') for k in v]
        ax.axvspan(-.5, .5, color='#ebebeb', zorder=0)
        ax.barh(y, v, color=cols_, height=.66, zorder=3)
        ax.axvline(0, color='#333', lw=.8, zorder=4)
        ax.set_yticks(y); ax.set_yticklabels([str(casos[i]) for i in o], fontsize=6.4)
        ax.set_ylabel('Caso', fontsize=7.5)
        ax.set_xlim(-1.15, 2.25); ax.set_xticks([-1, -.5, 0, .5, 1, 1.5, 2])
        ax.set_xticklabels(['−1,0', '−0,5', '0', '0,5', '1,0', '1,5', '2,0'], fontsize=6.8)
        ax.set_title(tit, fontsize=9, pad=5); ax.grid(axis='y', visible=False)
    fig.text(.5, -.03, 'Diferencia de calificación promedio (sistema − experto humano)',
             ha='center', fontsize=8)
    fig.text(.5, -.10, 'La banda gris marca el margen de equivalencia preespecificado de '
             '±0,5 puntos. Los casos se ordenan por magnitud dentro de cada panel.',
             ha='center', fontsize=6.4, style='italic', color='#555')
    fig.tight_layout(w_pad=2.5)
    fig.savefig(salida / 'figura3_diferencias.png', bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f'  figura3  n={len(casos)}  a favor sistema={sum(1 for x in rel if x>0)}')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--calculos', type=Path, required=True)
    ap.add_argument('--ciego', type=Path, required=True)
    ap.add_argument('--salida', type=Path, default=Path('figuras'))
    a = ap.parse_args()
    a.salida.mkdir(parents=True, exist_ok=True)

    print('Fuentes y huellas SHA-256:')
    for p in (a.calculos, a.ciego):
        print(f'  {p.name}: {sha256(p)}')
    print('Generando:')
    wbA = openpyxl.load_workbook(a.calculos, read_only=True, data_only=True)
    figura1(wbA, a.salida)
    figura2(wbA, a.salida)
    wbB = openpyxl.load_workbook(a.ciego, read_only=True, data_only=True)
    figura3(wbB, a.salida)
    print(f'Listo en {a.salida}/')


if __name__ == '__main__':
    main()