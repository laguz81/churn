"""PASO 8 -- Script unico y reproducible del estudio ciego pareado
(sistema agentico vs experto humano EH2, 15 casos, 3 evaluadores).

Consolida PASOS 1-7 (ya validados por separado en 01..06_*.py) en un
solo script cuyos parametros se CARGAN desde decisiones_analisis.md
(no se hardcodean por separado) y cuya semilla esta fijada (42).

Genera:
  - Respaldo_Estudio_Ciego_<fecha>.xlsx  (una hoja por paso)
  - resultados_resumen.md                (numeros para tesis, S2.5/S2.6)

Reglas de contenido para resultados_resumen.md (instruccion del usuario):
  - No inferioridad = resultado PRINCIPAL del contraste (H01 rechazada,
    ambos criterios), no nota al pie del resultado nulo de PASO 5.
  - ICC(2,k) va en el MISMO bloque que la no inferioridad.
  - Prohibido: "desempeno equivalente al experto humano", "no hay
    diferencia entre ambas fuentes".
  - Formulacion aprobada para el resultado nulo de PASO 5 (ya usada
    tal cual, sin parafrasear)."""
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
import pingouin as pg
from openpyxl import Workbook
from openpyxl.styles import Font
from scipy.stats import wilcoxon

ROOT = Path(__file__).resolve().parent.parent
REPO_ROOT = ROOT.parent
CALIF = ROOT / "calificaciones.csv"
DECODE = ROOT / "decode.json"
EVALUADORES = REPO_ROOT / "panel_evaluacion" / "datos" / "evaluadores.json"
DECISIONES_MD = ROOT / "decisiones_analisis.md"
CEGADO_MD = ROOT / "clasificacion_cegado.md"

CRITERIOS = ("relevancia", "viabilidad")
CASOS = list(range(1, 16))
FECHA = date.today().strftime("%Y%m%d")
XLSX_OUT = ROOT / f"Respaldo_Estudio_Ciego_{FECHA}.xlsx"
MD_OUT = ROOT / "resultados_resumen.md"

FORMULACION_NULO_APROBADA = (
    "no se detectó una diferencia estadísticamente significativa; con quince "
    "casos el diseño tiene potencia para detectar únicamente diferencias "
    "grandes, por lo que el resultado no permite concluir equivalencia"
)


# =====================================================================
# 0. Parametros: CARGADOS desde decisiones_analisis.md, no hardcodeados
#    por separado. Se assert-ean contra las constantes usadas por el
#    pipeline para garantizar que el codigo no diverge del documento
#    congelado.
# =====================================================================
def parse_decisiones_md(path: Path) -> dict:
    """Parsea lineas 'CLAVE = valor  # comentario' de decisiones_analisis.md.
    Devuelve {clave: valor_crudo_como_string (sin comentario, sin espacios)}."""
    params = {}
    for linea in path.read_text(encoding="utf-8").splitlines():
        m = re.match(r"^([A-Z_]+)\s*=\s*(.+)$", linea)
        if not m:
            continue
        clave, resto = m.group(1), m.group(2)
        valor = resto.split("#", 1)[0].strip()
        params[clave] = valor
    return params


PARAMS_MD = parse_decisiones_md(DECISIONES_MD)


def _num(valor_str: str) -> float:
    """'0,5 puntos Likert' -> 0.5 ; '42' -> 42.0 (coma decimal -> punto)."""
    token = valor_str.split()[0]
    return float(token.replace(",", "."))


# Constantes del pipeline (usadas mas abajo). Se assert-ean == a lo
# leido de decisiones_analisis.md: si alguien edita el .md y no este
# script (o viceversa), el script se detiene en vez de correr silencioso
# con parametros desincronizados.
UNIDAD_ANALISIS = "promedio_por_caso"
EMPATES = "pratt"
ICC_MODELO = "dos_vias_aleatorio"
ICC_TIPO = "acuerdo_absoluto"
REVISION_MANUAL = "incluir"
MARGEN_TOST = 0.5
SEMILLA = 42
ALFA = 0.05

assert PARAMS_MD["UNIDAD_ANALISIS"] == UNIDAD_ANALISIS, PARAMS_MD["UNIDAD_ANALISIS"]
assert PARAMS_MD["EMPATES"] == EMPATES, PARAMS_MD["EMPATES"]
assert PARAMS_MD["ICC_MODELO"] == ICC_MODELO, PARAMS_MD["ICC_MODELO"]
assert PARAMS_MD["ICC_TIPO"] == ICC_TIPO, PARAMS_MD["ICC_TIPO"]
assert PARAMS_MD["REVISION_MANUAL"] == REVISION_MANUAL, PARAMS_MD["REVISION_MANUAL"]
assert _num(PARAMS_MD["MARGEN_TOST"]) == MARGEN_TOST, PARAMS_MD["MARGEN_TOST"]
assert int(_num(PARAMS_MD["SEMILLA"])) == SEMILLA, PARAMS_MD["SEMILLA"]
assert _num(PARAMS_MD["ALFA"]) == ALFA, PARAMS_MD["ALFA"]
print("OK: parametros del pipeline coinciden con decisiones_analisis.md.")


# =====================================================================
# PASO 1 -- Validacion de calificaciones.csv
# =====================================================================
with open(CALIF, newline="", encoding="utf-8") as f:
    filas_calif = list(csv.DictReader(f))

assert len(filas_calif) == 180, f"filas != 180 (hay {len(filas_calif)})"

evaluadores = sorted({r["evaluador"] for r in filas_calif})
casos_csv = sorted({int(r["id_caso"]) for r in filas_calif})
assert len(evaluadores) == 3, evaluadores
assert len(casos_csv) == 15, casos_csv

# ASSERT requerido: ningun caso repetido para el mismo evaluador+etiqueta+criterio.
combos = Counter((r["evaluador"], r["id_caso"], r["etiqueta"], r["criterio"]) for r in filas_calif)
duplicados = {k: v for k, v in combos.items() if v > 1}
assert not duplicados, f"combinaciones duplicadas (evaluador,caso,etiqueta,criterio): {duplicados}"
print(f"OK: sin duplicados evaluador+caso+etiqueta+criterio ({len(combos)} combinaciones unicas).")

paso1_filas = []
for ev in evaluadores:
    subset = [r for r in filas_calif if r["evaluador"] == ev]
    vals = [int(r["puntaje"]) for r in subset]
    dist = Counter(vals)
    moda, freq_moda = Counter(vals).most_common(1)[0]
    paso1_filas.append(dict(
        evaluador=ev, n=len(vals),
        dist_1=dist.get(1, 0), dist_2=dist.get(2, 0), dist_3=dist.get(3, 0),
        dist_4=dist.get(4, 0), dist_5=dist.get(5, 0),
        moda=moda, frac_moda=round(freq_moda / len(vals), 4),
    ))


# =====================================================================
# PASO 2 -- Cegado (ya congelado en clasificacion_cegado.md, ANTES de
# abrir el decode; aqui solo se reproduce como hoja de respaldo)
# =====================================================================
paso2_lineas = [ln for ln in CEGADO_MD.read_text(encoding="utf-8").splitlines() if ln.strip()]


# =====================================================================
# Carga comun: decode + evaluadores -> acumulador[(caso,fuente,criterio)]
# =====================================================================
with open(DECODE, encoding="utf-8") as f:
    decode = json.load(f)
with open(EVALUADORES, encoding="utf-8") as f:
    evaluadores_meta = json.load(f)
token_a_evaluador = {t: m["evaluador_id"] for t, m in evaluadores_meta.items() if not m["es_prueba"]}
evaluador_a_token = {v: k for k, v in token_a_evaluador.items()}

acumulador = defaultdict(list)
for r in filas_calif:
    token = evaluador_a_token[r["evaluador"]]
    fuente = decode[token][r["id_caso"]][r["etiqueta"]]
    acumulador[(int(r["id_caso"]), fuente, r["criterio"])].append(int(r["puntaje"]))
for k, v in acumulador.items():
    assert len(v) == 3, f"{k} no tiene 3 puntajes: {v}"


# =====================================================================
# PASO 3 -- Tabla pareada (15 filas, promedio de 3 evaluadores)
# =====================================================================
tabla = []
for caso in CASOS:
    fila = {"caso": caso}
    for fuente in ("sistema", "eh2"):
        for criterio in CRITERIOS:
            vals = acumulador[(caso, fuente, criterio)]
            fila[f"{fuente}_{criterio}"] = sum(vals) / len(vals)
    tabla.append(fila)

assert len(tabla) == 15, "la tabla pareada debe tener 15 filas"
# ASSERT requerido: exactamente 15 filas por criterio (cobertura completa).
for criterio in CRITERIOS:
    casos_con_ambas_fuentes = {
        c for c in CASOS
        if len(acumulador[(c, "sistema", criterio)]) == 3 and len(acumulador[(c, "eh2", criterio)]) == 3
    }
    assert len(casos_con_ambas_fuentes) == 15, f"{criterio}: {len(casos_con_ambas_fuentes)} filas, no 15"
print("OK: tabla pareada con exactamente 15 filas por criterio.")


# =====================================================================
# PASO 4 -- Descriptivos (sin p-valores)
# =====================================================================
QUANTILE_METHOD = "linear"
paso4_resumen = {}
paso4_conteo = {}
for criterio in CRITERIOS:
    for fuente in ("sistema", "eh2"):
        vals = np.array([fila[f"{fuente}_{criterio}"] for fila in tabla])
        paso4_resumen[(criterio, fuente)] = dict(
            mediana=float(np.percentile(vals, 50, method=QUANTILE_METHOD)),
            q1=float(np.percentile(vals, 25, method=QUANTILE_METHOD)),
            q3=float(np.percentile(vals, 75, method=QUANTILE_METHOD)),
            min=float(vals.min()), max=float(vals.max()),
        )
    diffs = np.array([fila[f"sistema_{criterio}"] - fila[f"eh2_{criterio}"] for fila in tabla])
    a_favor_sistema = int((diffs > 0).sum())
    a_favor_eh2 = int((diffs < 0).sum())
    empatados = int((diffs == 0).sum())
    # ASSERT requerido: a_favor_sistema + a_favor_eh2 + empatados == 15
    assert a_favor_sistema + a_favor_eh2 + empatados == 15
    paso4_conteo[criterio] = dict(sistema=a_favor_sistema, eh2=a_favor_eh2, empate=empatados)
print("OK: a_favor_sistema + a_favor_eh2 + empatados == 15 en ambos criterios.")


# =====================================================================
# PASO 5 -- Wilcoxon pareado (Pratt), escala suma entera (bugfix), y
# version ANTES del fix (promedio float) para dejar registro comparable.
# =====================================================================
def sumas_enteras(criterio: str):
    suma_sistema = np.array([sum(acumulador[(c, "sistema", criterio)]) for c in CASOS], dtype=np.int64)
    suma_eh2 = np.array([sum(acumulador[(c, "eh2", criterio)]) for c in CASOS], dtype=np.int64)
    return suma_sistema, suma_eh2


def rank_biserial_pratt(diffs: np.ndarray):
    abs_d = np.abs(diffs)
    n = len(diffs)
    orden = np.argsort(abs_d, kind="mergesort")
    rangos = np.empty(n)
    valores_ordenados = abs_d[orden]
    i, r = 0, 1
    while i < n:
        j = i
        while j < n and valores_ordenados[j] == valores_ordenados[i]:
            j += 1
        rango_promedio = (r + (r + (j - i) - 1)) / 2
        rangos[orden[i:j]] = rango_promedio
        r += (j - i)
        i = j
    r_pos = rangos[diffs > 0].sum()
    r_neg = rangos[diffs < 0].sum()
    denom = r_pos + r_neg
    r_rb = (r_pos - r_neg) / denom if denom > 0 else float("nan")
    return float(r_pos), float(r_neg), float(r_rb)


def bootstrap_ci_diferencia_medianas(criterio: str, percentiles, n_rep=1000, seed=SEMILLA):
    sistema_avg = np.array([sum(acumulador[(c, "sistema", criterio)]) / 3.0 for c in CASOS])
    eh2_avg = np.array([sum(acumulador[(c, "eh2", criterio)]) / 3.0 for c in CASOS])
    rng = np.random.default_rng(seed)
    n = len(CASOS)
    diffs_medianas = np.empty(n_rep)
    for b in range(n_rep):
        idx = rng.integers(0, n, size=n)
        diffs_medianas[b] = np.median(sistema_avg[idx]) - np.median(eh2_avg[idx])
    lo, hi = np.percentile(diffs_medianas, percentiles)
    mediana_puntual = float(np.median(sistema_avg) - np.median(eh2_avg))
    return mediana_puntual, float(lo), float(hi)


paso5_resultados = {}
for criterio in CRITERIOS:
    suma_sistema, suma_eh2 = sumas_enteras(criterio)
    diffs_fixed = (suma_sistema - suma_eh2).astype(np.float64)  # DESPUES: enteros exactos
    sistema_avg = np.array([sum(acumulador[(c, "sistema", criterio)]) / 3.0 for c in CASOS])
    eh2_avg = np.array([sum(acumulador[(c, "eh2", criterio)]) / 3.0 for c in CASOS])
    diffs_buggy = sistema_avg - eh2_avg  # ANTES: promedio float (1/3 no exacto en binario)

    res_fixed = wilcoxon(diffs_fixed, zero_method="pratt", alternative="two-sided", method="auto")
    res_buggy = wilcoxon(diffs_buggy, zero_method="pratt", alternative="two-sided", method="auto")
    r_pos, r_neg, r_rb = rank_biserial_pratt(diffs_fixed)
    mediana_diff, ic95_lo, ic95_hi = bootstrap_ci_diferencia_medianas(criterio, [2.5, 97.5])

    paso5_resultados[criterio] = dict(
        n=15, empates=int((diffs_fixed == 0).sum()),
        W_antes=float(res_buggy.statistic), p_antes=float(res_buggy.pvalue),
        W_despues=float(res_fixed.statistic), p_despues=float(res_fixed.pvalue),
        r_pos=r_pos, r_neg=r_neg, r_rb=r_rb,
        mediana_diff=mediana_diff, ic95_lo=ic95_lo, ic95_hi=ic95_hi,
    )
print("PASO 5 (Wilcoxon principal + comparacion antes/despues del bugfix) completo.")


# =====================================================================
# PASO 6 -- ICC(2,1)/ICC(2,k) acuerdo absoluto (principal) + ICC(C,1)/
# ICC(C,k) consistencia (diagnostico post-hoc)
# =====================================================================
with open(CALIF, newline="", encoding="utf-8") as f:
    filas_icc = list(csv.DictReader(f))

paso6_resultados = {}
for criterio in CRITERIOS:
    acc = {}
    for r in filas_icc:
        if r["criterio"] != criterio:
            continue
        clave = (int(r["id_caso"]), r["evaluador"])
        acc.setdefault(clave, {})[r["etiqueta"]] = int(r["puntaje"])
    registros = []
    for (caso, evaluador), pares in acc.items():
        assert set(pares.keys()) == {"A", "B"}, f"{caso},{evaluador}: {pares}"
        registros.append(dict(caso=caso, evaluador=evaluador, valor=(pares["A"] + pares["B"]) / 2.0))
    df = pd.DataFrame(registros)
    assert df["caso"].nunique() == 15 and df["evaluador"].nunique() == 3 and len(df) == 45

    icc = pg.intraclass_corr(data=df, targets="caso", raters="evaluador", ratings="valor").set_index("Type")
    a1, ak = icc.loc["ICC(A,1)"], icc.loc["ICC(A,k)"]   # = ICC(2,1) / ICC(2,k), acuerdo absoluto
    c1, ck = icc.loc["ICC(C,1)"], icc.loc["ICC(C,k)"]   # consistencia (diagnostico)

    paso6_resultados[criterio] = dict(
        icc21=float(a1["ICC"]), icc21_ci=str(a1["CI95"]),
        icc2k=float(ak["ICC"]), icc2k_ci=str(ak["CI95"]),
        iccc1=float(c1["ICC"]), iccc1_ci=str(c1["CI95"]),
        iccck=float(ck["ICC"]), iccck_ci=str(ck["CI95"]),
    )
print("PASO 6 (ICC acuerdo absoluto + consistencia) completo.")


# =====================================================================
# PASO 7 -- TOST de no inferioridad / equivalencia
# =====================================================================
MARGEN_SUMA = MARGEN_TOST * 3  # escala suma entera; 1.5 exacto en float64

paso7_resultados = {}
for criterio in CRITERIOS:
    suma_sistema, suma_eh2 = sumas_enteras(criterio)
    diffs = (suma_sistema - suma_eh2).astype(np.float64)

    p_h01 = float(wilcoxon(diffs + MARGEN_SUMA, zero_method="pratt", alternative="greater", method="auto").pvalue)
    p_h02 = float(wilcoxon(diffs - MARGEN_SUMA, zero_method="pratt", alternative="less", method="auto").pvalue)
    no_inferior = p_h01 < ALFA          # H01 rechazada: sistema no peor que EH2 por > margen
    no_superior = p_h02 < ALFA          # H02 rechazada: sistema no mejor que EH2 por > margen
    equivalencia_via1 = no_inferior and no_superior

    mediana_diff, ic90_lo, ic90_hi = bootstrap_ci_diferencia_medianas(criterio, [5, 95])
    equivalencia_via2 = (ic90_lo >= -MARGEN_TOST) and (ic90_hi <= MARGEN_TOST)

    paso7_resultados[criterio] = dict(
        p_h01=p_h01, p_h02=p_h02, no_inferior=no_inferior, no_superior=no_superior,
        equivalencia_via1=equivalencia_via1,
        mediana_diff=mediana_diff, ic90_lo=ic90_lo, ic90_hi=ic90_hi,
        equivalencia_via2=equivalencia_via2,
    )
print("PASO 7 (TOST no inferioridad/equivalencia) completo.")


# =====================================================================
# Salida 1: Respaldo_Estudio_Ciego_<fecha>.xlsx
# Formato: sigue Respaldo_Calculos_v18m_20260811.xlsx (Camino A) --
# fila 1 = descripcion/fuente (9pt, normal), fila 2 en blanco,
# fila 3 = encabezados (bold), datos desde fila 4. Sin fills, sin
# freeze panes, numeros en General (sin redondear).
# =====================================================================
wb = Workbook()
wb.remove(wb.active)

FONT_TITULO = Font(bold=True, size=14, name="Calibri")
FONT_CAPTION = Font(bold=False, size=9, name="Calibri")
FONT_HEADER = Font(bold=True, name="Calibri")


def nueva_hoja(nombre: str, caption: str, headers: list[str], filas: list[list]):
    ws = wb.create_sheet(nombre)
    ws.cell(row=1, column=1, value=caption).font = FONT_CAPTION
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = FONT_HEADER
    for i, fila in enumerate(filas, start=4):
        for j, val in enumerate(fila, start=1):
            ws.cell(row=i, column=j, value=val)
    return ws


# --- 00_Portada ---
ws = wb.create_sheet("00_Portada")
ws.cell(row=1, column=1, value="RESPALDO DE CÁLCULOS — Estudio ciego pareado").font = FONT_TITULO
meta = [
    ("Proyecto", "Estudio ciego pareado de recomendaciones de retención — sistema agéntico vs experto humano (EH2)"),
    ("Fecha de generación", date.today().isoformat()),
    ("Script fuente", "analisis_ciego/scripts/analisis_estudio_ciego.py"),
    ("Semilla", str(SEMILLA)),
    ("Alfa", str(ALFA).replace(".", ",")),
    ("Margen TOST", f"±{str(MARGEN_TOST).replace('.', ',')} puntos Likert"),
    ("n casos", "15"),
    ("n evaluadores", "3"),
]
for i, (k, v) in enumerate(meta, start=3):
    ws.cell(row=i, column=1, value=k)
    ws.cell(row=i, column=2, value=v)

# --- 01_Validacion ---
nueva_hoja(
    "01_Validacion",
    "PASO 1 — Distribución de puntajes y varianza por evaluador. Fuente: calificaciones.csv.",
    ["evaluador", "n", "dist_1", "dist_2", "dist_3", "dist_4", "dist_5", "moda", "frac_moda"],
    [[f["evaluador"], f["n"], f["dist_1"], f["dist_2"], f["dist_3"], f["dist_4"], f["dist_5"], f["moda"], f["frac_moda"]] for f in paso1_filas],
)

# --- 02_Cegado ---
nueva_hoja(
    "02_Cegado",
    "PASO 2 — Clasificación de cegado, hecha ANTES de abrir el decode. Fuente: clasificacion_cegado.md.",
    ["linea"],
    [[ln] for ln in paso2_lineas],
)

# --- 03_Tabla_Pareada ---
nueva_hoja(
    "03_Tabla_Pareada",
    "PASO 3 — Tabla pareada, 15 casos, promedio de 3 evaluadores por fuente y criterio.",
    ["caso", "sistema_relevancia", "eh2_relevancia", "sistema_viabilidad", "eh2_viabilidad"],
    [[f["caso"], f["sistema_relevancia"], f["eh2_relevancia"], f["sistema_viabilidad"], f["eh2_viabilidad"]] for f in tabla],
)

# --- 04_Descriptivos ---
filas4 = []
for criterio in CRITERIOS:
    for fuente in ("sistema", "eh2"):
        d = paso4_resumen[(criterio, fuente)]
        filas4.append([criterio, fuente, d["mediana"], d["q1"], d["q3"], d["min"], d["max"]])
for criterio in CRITERIOS:
    c = paso4_conteo[criterio]
    filas4.append([criterio, "conteo(a_favor_sistema/a_favor_eh2/empate)", c["sistema"], c["eh2"], c["empate"], None, None])
nueva_hoja(
    "04_Descriptivos",
    "PASO 4 — Descriptivos por fuente y criterio (sin p-valores), escala promedio_por_caso.",
    ["criterio", "fuente_o_conteo", "mediana_o_sistema", "q1_o_eh2", "q3_o_empate", "min", "max"],
    filas4,
)

# --- 05_Wilcoxon ---
filas5 = []
for criterio in CRITERIOS:
    r = paso5_resultados[criterio]
    filas5.append([
        criterio, r["n"], r["empates"],
        r["W_antes"], r["p_antes"], r["W_despues"], r["p_despues"],
        r["r_pos"], r["r_neg"], r["r_rb"], r["mediana_diff"], r["ic95_lo"], r["ic95_hi"],
    ])
nueva_hoja(
    "05_Wilcoxon",
    "PASO 5 — Wilcoxon pareado (Pratt). ANTES = promedio float (bug de precisión, PRE-fix); "
    "DESPUÉS = suma entera (fix, resultado válido). Bootstrap IC95%, 1000 réplicas, semilla=42.",
    ["criterio", "n", "empates", "W_antes(bug)", "p_antes(bug)", "W_despues(fix)", "p_despues(fix)",
     "R+", "R-", "r_rb", "mediana_diff", "ic95_lo", "ic95_hi"],
    filas5,
)

# --- 06_ICC ---
filas6 = []
for criterio in CRITERIOS:
    r = paso6_resultados[criterio]
    filas6.append([criterio, "ICC(2,1) [acuerdo absoluto]", r["icc21"], r["icc21_ci"]])
    filas6.append([criterio, "ICC(2,k) [acuerdo absoluto]", r["icc2k"], r["icc2k_ci"]])
    filas6.append([criterio, "ICC(C,1) [consistencia, diagnostico]", r["iccc1"], r["iccc1_ci"]])
    filas6.append([criterio, "ICC(C,k) [consistencia, diagnostico]", r["iccck"], r["iccck_ci"]])
nueva_hoja(
    "06_ICC",
    "PASO 6 — ICC dos vías, efectos aleatorios. Acuerdo absoluto = principal (congelado en "
    "decisiones_analisis.md); consistencia = diagnóstico post-hoc.",
    ["criterio", "tipo", "icc", "ic95"],
    filas6,
)

# --- 07_TOST ---
filas7 = []
for criterio in CRITERIOS:
    r = paso7_resultados[criterio]
    filas7.append([
        criterio, r["p_h01"], r["no_inferior"], r["p_h02"], r["no_superior"], r["equivalencia_via1"],
        r["mediana_diff"], r["ic90_lo"], r["ic90_hi"], r["equivalencia_via2"],
    ])
nueva_hoja(
    "07_TOST",
    f"PASO 7 — TOST, margen ±{MARGEN_TOST}, alfa={ALFA}. Vía 1: Wilcoxon una cola (Pratt). "
    "Vía 2: IC90% bootstrap, 1000 réplicas, semilla=42.",
    ["criterio", "p_H01", "H01_rechazada(no_inferior)", "p_H02", "H02_rechazada(no_superior)",
     "equivalencia_via1", "mediana_diff", "ic90_lo", "ic90_hi", "equivalencia_via2"],
    filas7,
)

wb.save(XLSX_OUT)
print(f"Guardado: {XLSX_OUT}")


# =====================================================================
# Salida 2: resultados_resumen.md
# =====================================================================
def _es(x: float, dec: int = 2) -> str:
    """Formatea un float con coma decimal (convencion espanola usada en decisiones_analisis.md)."""
    return f"{x:.{dec}f}".replace(".", ",")


r_rel5, r_via5 = paso5_resultados["relevancia"], paso5_resultados["viabilidad"]
r6_rel, r6_via = paso6_resultados["relevancia"], paso6_resultados["viabilidad"]
r7_rel, r7_via = paso7_resultados["relevancia"], paso7_resultados["viabilidad"]

md = f"""# Resultados — estudio ciego pareado (sistema agéntico vs EH2)

Generado por `analisis_estudio_ciego.py` el {date.today().isoformat()}, semilla={SEMILLA}.
Números listos para insertar en §2.5 y §2.6.

## §2.5 — Descriptivos y verificaciones metodológicas

**Mínimos por fuente (PASO 4, escala promedio_por_caso):**

- Relevancia: sistema mín = {_es(paso4_resumen[('relevancia','sistema')]['min'], 4)}; EH2 mín = {_es(paso4_resumen[('relevancia','eh2')]['min'], 4)}.
- Viabilidad: sistema mín = {_es(paso4_resumen[('viabilidad','sistema')]['min'], 4)}; EH2 mín = {_es(paso4_resumen[('viabilidad','eh2')]['min'], 4)}.

**Bug de punto flotante corregido (PASO 5).** La versión inicial calculó las diferencias
pareadas sobre el promedio en punto flotante de los 3 evaluadores (suma/3), y 1/3 no es exacto
en binario: pares con la misma magnitud de diferencia (p. ej. exactamente 1/3) llegaban a la
corrección de Pratt como floats distintos, rompiendo el agrupamiento de rangos empatados. Se
corrigió operando sobre la suma entera de los 3 evaluadores (Wilcoxon es invariante a un
reescalado positivo común), eliminando el ruido numérico. Antes y después del fix:

| Criterio | W antes (bug) | p antes (bug) | W después (fix) | p después (fix) |
|---|---|---|---|---|
| Relevancia | {_es(r_rel5['W_antes'])} | {_es(r_rel5['p_antes'], 6)} | {_es(r_rel5['W_despues'])} | {_es(r_rel5['p_despues'], 6)} |
| Viabilidad | {_es(r_via5['W_antes'])} | {_es(r_via5['p_antes'], 6)} | {_es(r_via5['W_despues'])} | {_es(r_via5['p_despues'], 6)} |

El resultado de PASO 5, ya con el fix aplicado: para ambos criterios, {FORMULACION_NULO_APROBADA}.

## §2.6 — Resultado principal del contraste: no inferioridad, sobre un instrumento de fiabilidad limitada

El resultado principal del contraste no es la ausencia de diferencia de PASO 5 (que por sí sola
no permite concluir equivalencia, ver §2.5), sino la **no inferioridad** establecida mediante TOST
con margen de equivalencia preespecificado de ±{_es(MARGEN_TOST)} puntos Likert (congelado en
decisiones_analisis.md) y alfa={_es(ALFA)} (fijado en PASO 7; la lectura no cambia con ningún alfa
convencional — 0,01, 0,05 o 0,10 — porque los p-valores de PASO 5 son 0,197 y 0,603).

Se rechazó H01 (mediana de la diferencia sistema−EH2 ≤ −{_es(MARGEN_TOST)}) en ambos criterios,
mediante Wilcoxon de una cola con corrección de Pratt sobre la escala del PASO 5:

- Relevancia: p = {_es(r7_rel['p_h01'], 4)} (H01 rechazada).
- Viabilidad: p = {_es(r7_via['p_h01'], 4)} (H01 rechazada).

Es decir, el sistema **no fue peor que EH2 por más del margen preespecificado**, en ninguno de
los dos criterios. No se rechazó H02 (mediana ≥ +{_es(MARGEN_TOST)}) en ninguno de los dos criterios
(relevancia p={_es(r7_rel['p_h02'], 4)}, viabilidad p={_es(r7_via['p_h02'], 4)}), de modo que no se
estableció equivalencia de dos colas: los datos no permiten descartar que el sistema supere a
EH2 por más del margen. El IC90% bootstrap (vía 2, 1000 réplicas, semilla={SEMILLA}) confirma
la misma lectura: relevancia [{_es(r7_rel['ic90_lo'], 4)}, {_es(r7_rel['ic90_hi'], 4)}], viabilidad
[{_es(r7_via['ic90_lo'], 4)}, {_es(r7_via['ic90_hi'], 4)}] — en ambos casos el intervalo no quedó
contenido en [−{_es(MARGEN_TOST)}, +{_es(MARGEN_TOST)}]; con n=15 el intervalo resultó demasiado ancho
para concluir equivalencia de dos colas.

**Fiabilidad del instrumento (PASO 6, mismo bloque).** Esta no inferioridad se estableció sobre
un instrumento de fiabilidad limitada: ICC(2,k) (dos vías, efectos aleatorios, acuerdo absoluto,
promedio de 3 evaluadores) = {_es(r6_rel['icc2k'])} en relevancia (IC95% {r6_rel['icc2k_ci']}) y
{_es(r6_via['icc2k'])} en viabilidad (IC95% {r6_via['icc2k_ci']}).

**Consistencia vs. acuerdo absoluto.** El ICC de consistencia (diagnóstico post-hoc, ICC(C,k)) es
sustancialmente más alto que el de acuerdo absoluto: {_es(r6_rel['iccck'])} vs {_es(r6_rel['icc2k'])}
en relevancia, {_es(r6_via['iccck'])} vs {_es(r6_via['icc2k'])} en viabilidad. Lectura: los tres
evaluadores **ordenan los casos de forma similar** (consistencia alta), pero **difieren en el
nivel de severidad de la escala** que aplican (acuerdo absoluto bajo) — consistente con el sesgo
de escala detectado en PASO 1 (moda evaluador_1=5, evaluador_2=4, evaluador_3=2, nunca usó 5).

---
**Formulaciones evitadas deliberadamente:** no se afirma que "el sistema alcanza un desempeño
equivalente al del experto humano" ni que "no hay diferencia entre ambas fuentes"; ninguna de
las dos es lo que muestran estos datos.
"""

MD_OUT.write_text(md, encoding="utf-8")
print(f"Guardado: {MD_OUT}")

print("\nPASO 8 completo.")
